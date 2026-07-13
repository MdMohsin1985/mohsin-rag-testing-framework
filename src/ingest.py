"""Ingest TXT, PDF, and Excel documents into a persistent ChromaDB collection."""

from __future__ import annotations

import hashlib
import os
from pathlib import Path
from typing import Iterable

import chromadb
from chromadb.utils.embedding_functions import OpenAIEmbeddingFunction
from dotenv import load_dotenv
from openpyxl import load_workbook
from pypdf import PdfReader


ROOT_DIR = Path(__file__).resolve().parents[1]
DOCUMENTS_DIR = ROOT_DIR / "documents"
CHROMA_DIR = ROOT_DIR / "chroma_db"
COLLECTION_NAME = "rag_documents"
EMBEDDING_MODEL = "text-embedding-3-small"
SUPPORTED_EXTENSIONS = {".txt", ".pdf", ".xlsx", ".xlsm", ".xltx", ".xltm"}


def require_api_key() -> str:
    """Load and validate the OpenAI API key used for embeddings."""
    load_dotenv(ROOT_DIR / ".env")
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        raise RuntimeError(
            "OPENAI_API_KEY is missing. Create a .env file from .env.example first."
        )
    return api_key


def read_txt(path: Path) -> str:
    """Read a plain text document."""
    return path.read_text(encoding="utf-8").strip()


def read_pdf(path: Path) -> str:
    """Extract text from every page of a PDF document."""
    reader = PdfReader(str(path))
    pages: list[str] = []
    for page_number, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""
        if text.strip():
            pages.append(f"Page {page_number}:\n{text.strip()}")
    return "\n\n".join(pages).strip()


def read_excel(path: Path) -> str:
    """Extract visible cell values from each sheet in an Excel workbook."""
    workbook = load_workbook(filename=path, read_only=True, data_only=True)
    sections: list[str] = []

    for sheet in workbook.worksheets:
        rows: list[str] = []
        for row in sheet.iter_rows(values_only=True):
            values = [str(value).strip() for value in row if value is not None]
            if values:
                rows.append(" | ".join(values))

        if rows:
            sections.append(f"Sheet: {sheet.title}\n" + "\n".join(rows))

    workbook.close()
    return "\n\n".join(sections).strip()


def read_document(path: Path) -> str:
    """Route supported document types to the right text extractor."""
    suffix = path.suffix.lower()
    if suffix == ".txt":
        return read_txt(path)
    if suffix == ".pdf":
        return read_pdf(path)
    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return read_excel(path)
    raise ValueError(f"Unsupported file type: {path.suffix}")


def read_documents(documents_dir: Path = DOCUMENTS_DIR) -> Iterable[tuple[Path, str]]:
    """Yield every non-empty supported document in the documents folder."""
    documents_dir.mkdir(parents=True, exist_ok=True)
    for path in sorted(documents_dir.iterdir()):
        if not path.is_file() or path.suffix.lower() not in SUPPORTED_EXTENSIONS:
            continue
        try:
            text = read_document(path)
        except Exception as exc:
            print(f"Skipping {path.name}: {exc}")
            continue
        if text:
            yield path, text


def chunk_text(text: str, chunk_size: int = 800, overlap: int = 120) -> list[str]:
    """Split text into overlapping character chunks."""
    if chunk_size <= overlap:
        raise ValueError("chunk_size must be greater than overlap")

    chunks: list[str] = []
    start = 0
    while start < len(text):
        end = start + chunk_size
        chunk = text[start:end].strip()
        if chunk:
            chunks.append(chunk)
        if end >= len(text):
            break
        start += chunk_size - overlap
    return chunks


def get_collection():
    """Create or load the ChromaDB collection with OpenAI embeddings."""
    api_key = require_api_key()
    client = chromadb.PersistentClient(path=str(CHROMA_DIR))
    embedding_function = OpenAIEmbeddingFunction(
        api_key=api_key,
        model_name=EMBEDDING_MODEL,
    )
    return client.get_or_create_collection(
        name=COLLECTION_NAME,
        embedding_function=embedding_function,
        metadata={"hnsw:space": "cosine"},
    )


def build_chunk_id(path: Path, index: int, chunk: str) -> str:
    """Build a stable ID so re-running ingestion updates existing chunks."""
    digest = hashlib.sha1(chunk.encode("utf-8")).hexdigest()[:12]
    return f"{path.stem}-{index}-{digest}"


def ingest_documents() -> int:
    """Read supported documents, split them, and upsert chunks into ChromaDB."""
    collection = get_collection()
    ids: list[str] = []
    chunks: list[str] = []
    metadatas: list[dict[str, str | int]] = []
    sources: set[str] = set()

    for path, text in read_documents():
        sources.add(path.name)
        for index, chunk in enumerate(chunk_text(text)):
            ids.append(build_chunk_id(path, index, chunk))
            chunks.append(chunk)
            metadatas.append(
                {
                    "source": path.name,
                    "file_type": path.suffix.lower(),
                    "chunk_index": index,
                }
            )

    if not chunks:
        extensions = ", ".join(sorted(SUPPORTED_EXTENSIONS))
        print(f"No supported documents found in {DOCUMENTS_DIR}. Supported: {extensions}")
        return 0

    existing = collection.get(include=["metadatas"])
    existing_sources = {
        metadata["source"]
        for metadata in existing.get("metadatas", [])
        if metadata and metadata.get("source")
    }
    for stale_source in existing_sources - sources:
        collection.delete(where={"source": stale_source})

    # Replace chunks for each current source so changed files cannot leave stale chunks behind.
    for source in sources:
        collection.delete(where={"source": source})

    collection.upsert(ids=ids, documents=chunks, metadatas=metadatas)
    print(f"Ingested {len(chunks)} chunks from {len(set(m['source'] for m in metadatas))} files.")
    return len(chunks)


if __name__ == "__main__":
    try:
        ingest_documents()
    except RuntimeError as exc:
        raise SystemExit(f"Error: {exc}") from exc
